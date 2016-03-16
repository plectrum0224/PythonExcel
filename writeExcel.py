#!/usr/bin/env python
# encoding: utf-8


"""
@version: ??
@author: phpergao
@license: Apache Licence 
@contact: endoffight@gmail.com
@site: http://
@software: PyCharm
@file: writeExcel.py
@time: 2016/3/15 22:40
"""


import openpyxl

# wb = openpyxl.load_workbook('example.xlsx')
#
# sheet = wb.get_active_sheet()
# sheet.title = 'Spam Spam Spam'
#
# wb.save('example_copy.xlsx')

# wb = openpyxl.Workbook()
# print(wb.get_sheet_names())
# wb.create_sheet()
# print(wb.get_sheet_names())
# wb.create_sheet(index=0, title='First Sheet')
# print(wb.get_sheet_names())
# wb.create_sheet(index=2, title='Middle Sheet')
# print(wb.get_sheet_names())
# wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
# wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
# print(wb.get_sheet_names())

# wb = openpyxl.load_workbook('produceSales.xlsx')
# sheet = wb.get_sheet_by_name('Sheet')
# PRICE_UPDATES = {
#     'Garlic': 3.07,
#     'Celery': 1.19,
#     'Lemon':1.27
# }
#
# for rowNum in range(2, sheet.get_highest_row()):
#     produceName = sheet.cell(row=rowNum, column=1).value
#     if produceName in PRICE_UPDATES:
#         sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
#
# wb.save('updatedProduceSales.xlsx')


listNum = [1, 2, 3, 100, 22,3,33,2,2,1,23,4,4,5,5]
def recordList(listType, i):
    countNum = 0
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Signals')
    sheet = wb.get_sheet_by_name('Signals')
    for columnNum in range(1,i):
        for rowNum in range(1, len(listNum)+1):
            sheet.cell(row=rowNum, column=columnNum).value = listNum[rowNum-1]
        countNum += 1
    if countNum == (i-1):
        wb.save('SignalExcel.xlsx')

recordList(listNum, 10)